import pdfkit
from jinja2 import Environment, FileSystemLoader
import os
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill, Font

def generate_pdf(context, output_path):
    template_dir = os.path.dirname(os.path.abspath('./template/template.html'))
    template_file = os.path.basename('./template.html')
    
    env = Environment(loader=FileSystemLoader(template_dir))
    template = env.get_template(template_file)
    
    html_out = template.render(context)
    css_abs_path = os.path.abspath('./template/template.css')
    # PDFKIT_CONFIG = pdfkit.configuration(wkhtmltopdf='/app/bin/wkhtmltopdf')
    # PDFKIT_CONFIG = pdfkit.configuration(wkhtmltopdf="C:/Program Files/wkhtmltopdf/bin/wkhtmltopdf.exe")
    PDFKIT_CONFIG = pdfkit.configuration(wkhtmltopdf="/usr/bin/wkhtmltopdf")

    
    pdfkit.from_string(html_out, output_path, options={"enable-local-file-access": ""}, css=css_abs_path, configuration=PDFKIT_CONFIG)





def create_excel_report_with_charts(data):
    # Crie um novo livro do Excel
    workbook = openpyxl.Workbook()
    
    # Obtenha a planilha ativa
    worksheet = workbook.active
    worksheet.title = "Relatório Geral"
    
    # Escreva os cabeçalhos na primeira planilha
    headers = ["Nome do Aluno", "Turma", "Número de Visitas", "Número de Ligações", "Número de Atendimentos", "Status do Caso", "Urgência do Caso", "Faltas"]
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        cell.font = Font(bold=True)
    
    # Escreva os dados na primeira planilha
    for row, caso in enumerate(data, 2):
        worksheet.cell(row=row, column=1, value=caso["aluno"]["nome"])
        worksheet.cell(row=row, column=2, value=caso["aluno"]["turma"])
        worksheet.cell(row=row, column=3, value=caso["n_visita"])
        worksheet.cell(row=row, column=4, value=caso["n_ligacao"])
        worksheet.cell(row=row, column=5, value=caso["n_atendimento"])
        worksheet.cell(row=row, column=6, value=caso["status"])
        worksheet.cell(row=row, column=7, value=caso["urgencia"])
        worksheet.cell(row=row, column=8, value=caso["faltas"])
    
    # Ajuste a largura das colunas
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width
    
    # Crie uma nova planilha para os gráficos
    chart_worksheet = workbook.create_sheet("Gráficos")
    
    # Obtenha os dados para os gráficos
    nomes_alunos = [caso["aluno"]["nome"] for caso in data]
    n_visitas = [caso["n_visita"] for caso in data]
    n_ligacoes = [caso["n_ligacao"] for caso in data]
    n_atendimentos = [caso["n_atendimento"] for caso in data]
    
    # Crie o gráfico de barras para alunos x visitas
    visitas_chart = BarChart()
    visitas_chart.type = "col"
    visitas_chart.style = 10
    visitas_chart.title = "Alunos x Visitas"
    visitas_chart.x_axis.title = "Alunos"
    visitas_chart.y_axis.title = "Número de Visitas"
    
    visitas_values = Reference(chart_worksheet, min_row=1, min_col=1, max_row=len(nomes_alunos), max_col=1)
    visitas_chart.add_data(visitas_values, from_rows=True, titles_from_data=True)
    chart_worksheet.add_chart(visitas_chart, "A1")
    
    # Crie o gráfico de barras para alunos x ligações
    ligacoes_chart = BarChart()
    ligacoes_chart.type = "col"
    ligacoes_chart.style = 11
    ligacoes_chart.title = "Alunos x Ligações"
    ligacoes_chart.x_axis.title = "Alunos"
    ligacoes_chart.y_axis.title = "Número de Ligações"
    
    ligacoes_values = Reference(chart_worksheet, min_row=1, min_col=2, max_row=len(nomes_alunos), max_col=2)
    ligacoes_chart.add_data(ligacoes_values, from_rows=True, titles_from_data=True)
    chart_worksheet.add_chart(ligacoes_chart, "A16")
    
    # Crie o gráfico de barras para alunos x atendimentos
    atendimentos_chart = BarChart()
    atendimentos_chart.type = "col"
    atendimentos_chart.style = 12
    atendimentos_chart.title = "Alunos x Atendimentos"
    atendimentos_chart.x_axis.title = "Alunos"
    atendimentos_chart.y_axis.title = "Número de Atendimentos"
    
    atendimentos_values = Reference(chart_worksheet, min_row=1, min_col=3, max_row=len(nomes_alunos), max_col=3)
    atendimentos_chart.add_data(atendimentos_values, from_rows=True, titles_from_data=True)
    chart_worksheet.add_chart(atendimentos_chart, "A31")
    
    return workbook